from django.shortcuts import render
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from openpyxl import load_workbook
import string
import locale
from io import BytesIO
from openpyxl.writer.excel import save_virtual_workbook



def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num

# Create your views here.

def final(request):
    return render(request,'final.html')



def index(request):
    headers=[]
    if request.method == "POST" and request.FILES['myfile']:
        filepath = request.FILES['myfile']
        wb=load_workbook(filepath)
        wb.create_sheet('newOutput')

        input_sheet = wb['Input']
        output_sheet = wb['newOutput']
        rule_sheet = wb['Rule']

        # type1 = request.POST.get('product_type')
        # if(type1 == "grocery"):
        #     headers[0] = ""
        #     headers[1] = ""
        #     headers[2] = ""
        #     headers[3] = ""
        #     headers[4] = ""
        # elif(type1 == ""):
        #     headers[0] = ""
        #     headers[1] = ""
        #     headers[2] = ""
        #     headers[3] = ""
        #     headers[4] = ""





        for row in rule_sheet.iter_rows('A2:D35'):
        	out = row[0].value
        	inp = row[1].value
        	rule = row[2].value
        	value1 = row[3].value

        	#For copy --> 1
        	if(rule == 1):
        		for i in range(3,1000):
        			if(input_sheet.cell(row = i, column = col2num(inp)).value != None):
        				output_sheet.cell(row = i, column = col2num(out)).value = input_sheet.cell(row = i, column = col2num(inp)).value

        	#For concat --> 2
        	elif(rule == 2):

        		my_data = inp.split(',')

        		for i in range(3,1000):
        			for j in my_data:

        				if(input_sheet.cell(row = i, column = col2num(j)).value != None):
        					fin = ""

        					for a in my_data:
        						fin = fin +" "+str(input_sheet.cell(row = i, column = col2num(a)).value)
        					output_sheet.cell(row = i, column = col2num(out)).value = fin











        	#For images --> 3
        	elif(rule == 3):
        		for i in range(3,1000):
        			if(input_sheet.cell(row = i, column = col2num(inp)).value != None):
        				output_sheet.cell(row = i, column = col2num(out)).hyperlink = value1+input_sheet.cell(row = i, column = col2num(inp)).value+".jpg"




            # #For Bullet Points --> 4
        	# elif(rule == 4):
            #
        	# 	my_data = inp.split(',')
            #
        	# 	for i in range(3,1000):
        	# 		for j in my_data:
            #
        	# 			if(input_sheet.cell(row = i, column = col2num(j)).value != None):
        	# 				fin = ""
            #
        	# 				for a,a1 in zip(my_data,range(0,5)):
        	# 					fin = fin +" "+headers[a1]+": "+str(input_sheet.cell(row = i, column = col2num(a)).value)+";"
        	# 				output_sheet.cell(row = i, column = col2num(out)).value = fin







































        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="final_report_flat.xlsx"'
        return response



    else:
        return render(request,'index.html')
    return render(request,'index.html')
